"""Deep, detailed catalog of Data/: full columns, coverage (year/country ranges),
key numeric ranges, archive/Excel internals, and duplicate detection."""
import gzip, hashlib, io, json, os, zipfile
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[2] / "Data"
OUT = Path(__file__).resolve().parents[2] / "results" / "audit" / "data_folder_catalog_deep.json"

YEAR_HINTS = ["year", "reportingyear", "ef year", "time_period", "yearvalid"]
CTRY_HINTS = ["country", "country_code", "countryname", "geo", "country_id", "registry_id", "region"]
EMIS_HINTS = ["verified", "releases", "co2", "ghg", "emission", "carbon", "intensity", "value", "footprint", "amount"]


def cov_from_df(df):
    info = {}
    low = {c.lower(): c for c in df.columns}
    for h in YEAR_HINTS:
        if h in low:
            s = pd.to_numeric(df[low[h]], errors="coerce").dropna()
            s = s[(s > 1900) & (s < 2100)]
            if len(s):
                info["year_range"] = [int(s.min()), int(s.max())]
            break
    for h in CTRY_HINTS:
        if h in low:
            u = df[low[h]].dropna().astype(str).unique()
            if 1 < len(u) <= 300:
                info["countries_n"] = int(len(u))
                info["countries_sample"] = sorted(u)[:30]
            break
    for c in df.columns:
        if any(h in c.lower() for h in EMIS_HINTS):
            s = pd.to_numeric(df[c], errors="coerce").dropna()
            if len(s) > 5 and s.abs().sum() > 0:
                info.setdefault("numeric_cols", {})[c] = {
                    "min": round(float(s.min()), 4), "median": round(float(s.median()), 4),
                    "max": round(float(s.max()), 4), "nonzero_frac": round(float((s != 0).mean()), 3)}
    return info


def profile_csv_bytes(b, name, nrows=200000):
    try:
        head = pd.read_csv(io.BytesIO(b), nrows=5, low_memory=False)
        cols = list(head.columns)
        df = pd.read_csv(io.BytesIO(b), nrows=nrows, low_memory=False)
        d = {"kind": "csv", "n_cols": len(cols), "columns": cols, "sampled_rows": len(df)}
        d.update(cov_from_df(df))
        return d
    except Exception as e:
        return {"kind": "csv", "error": str(e)[:100]}


def profile(p: Path):
    suf = p.suffix.lower()
    try:
        if suf == ".csv":
            with open(p, "rb") as f:
                b = f.read(40_000_000)
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                rows = sum(1 for _ in f) - 1
            d = profile_csv_bytes(b, p.name)
            d["rows_total"] = rows
            return d
        if suf == ".parquet":
            import pyarrow.parquet as pq
            md = pq.read_metadata(p)
            cols = [md.schema.column(i).name for i in range(md.num_columns)]
            d = {"kind": "parquet", "rows": md.num_rows, "n_cols": len(cols), "columns": cols}
            try:
                df = pq.read_table(p, columns=cols[: min(len(cols), 60)]).slice(0, 150000).to_pandas()
                d.update(cov_from_df(df))
            except Exception:
                pass
            return d
        if suf == ".json":
            d0 = json.loads(p.read_text(encoding="utf-8", errors="replace"))
            if isinstance(d0, list):
                return {"kind": "json-list", "len": len(d0),
                        "item_keys": list(d0[0].keys())[:40] if d0 and isinstance(d0[0], dict) else None}
            if isinstance(d0, dict):
                return {"kind": "json-dict", "keys": list(d0.keys())[:60]}
            return {"kind": "json"}
        if suf == ".zip":
            with zipfile.ZipFile(p) as z:
                members = [{"name": i.filename, "size": i.file_size} for i in z.infolist()]
            return {"kind": "zip", "n_members": len(members), "members": members}
        if suf == ".gz":
            with gzip.open(p, "rt", encoding="utf-8", errors="replace") as f:
                header = f.readline().strip()
            sep = "\t" if "\t" in header else ","
            return {"kind": "gzip-table", "columns": header.split(sep)[:40]}
        if suf in (".tsv", ".txt"):
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                header = f.readline().strip()
                rows = sum(1 for _ in f)
            sep = "\t" if "\t" in header else ","
            return {"kind": "tsv", "rows_total": rows, "columns": header.split(sep)[:40]}
        if suf in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True)
            sheets = {}
            for sn in wb.sheetnames[:12]:
                ws = wb[sn]
                hdr = next(ws.iter_rows(max_row=1, values_only=True), ())
                sheets[sn] = {"n_rows_approx": ws.max_row,
                              "columns": [str(c) for c in hdr][:40] if hdr else []}
            wb.close()
            return {"kind": "excel", "sheets": sheets}
        if suf in (".accdb", ".mdb"):
            return {"kind": "ms-access-db", "note": "MS Access; open with mdbtools/pyodbc"}
        if suf == ".pdf":
            return {"kind": "pdf"}
        return {"kind": suf or "other"}
    except Exception as e:
        return {"kind": suf, "error": str(e)[:100]}


def main():
    files = []
    for dp, _, fns in os.walk(ROOT):
        for f in fns:
            p = Path(dp) / f
            try:
                sz = p.stat().st_size
            except Exception:
                sz = 0
            files.append((sz, p))
    files.sort(key=lambda x: -x[0])

    # duplicate detection by (size) then quick hash of first 1MB for same-size groups
    bysize = {}
    for sz, p in files:
        bysize.setdefault(sz, []).append(p)
    dups = {}
    for sz, ps in bysize.items():
        if len(ps) > 1 and sz > 1000:
            for p in ps:
                try:
                    h = hashlib.md5(open(p, "rb").read(1_000_000)).hexdigest()[:10]
                except Exception:
                    h = "err"
                dups.setdefault(f"{sz}:{h}", []).append(str(p.relative_to(ROOT)).replace("\\", "/"))
    dup_groups = {k: v for k, v in dups.items() if len(v) > 1}

    cat = {"root": str(ROOT), "total_files": len(files),
           "total_gb": round(sum(s for s, _ in files) / 1e9, 3),
           "duplicate_groups": dup_groups, "files": []}
    for sz, p in files:
        rel = str(p.relative_to(ROOT)).replace("\\", "/")
        e = {"path": rel, "size": sz, "size_h": f"{sz/1e6:.1f}MB" if sz >= 1e6 else f"{sz/1e3:.0f}KB"}
        # profile everything; skip body-read only for the multi-GB opaque blobs
        if sz > 1_500_000_000 and p.suffix.lower() not in (".parquet", ".zip"):
            e["kind"] = "large-blob-" + (p.suffix.lower() or "")
        else:
            e.update(profile(p))
        cat["files"].append(e)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(cat, indent=1, default=str), encoding="utf-8")
    print(f"{cat['total_files']} files, {cat['total_gb']} GB, "
          f"{len(dup_groups)} duplicate groups -> {OUT}")


if __name__ == "__main__":
    main()
